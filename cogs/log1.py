import discord
from discord import ui
from discord.ui import Button, View
from discord.ext import commands, tasks
from discord.ext.commands import has_permissions, CheckFailure, HybridCommand
import time
import os
import pymongo
from datetime import datetime
import asyncio
from asyncio import sleep
import json 
import socket
import discord.utils
import requests
from pydub import AudioSegment
import urllib.request
import re
from urllib.parse import urlparse
from discord import Embed
import openpyxl



class Log1(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

    async def read_log(self, ctx: commands.Context, file_name: str):
        if not ctx.author.guild_permissions.administrator:
            await ctx.channel.send('권한이 없습니다.', delete_after=1)
            return

        folder_path = f"./dat/{ctx.guild.id}"
        files = os.listdir(folder_path)

        new_file_name = None
        for file in files:
            name, extension = os.path.splitext(file)
            if name == file_name:
                new_file_name = name + extension
                break
        if new_file_name is None:
            await ctx.send(f'파일 {file_name}이 존재하지 않습니다.\n!파일보기 로 확인하세요')
            return
        
        wb = openpyxl.load_workbook(f'./dat/{ctx.guild.id}/{new_file_name}')
        ws = wb.worksheets[0]
        sheet = wb.worksheets[0]

        max_column = sheet.max_column
        max_row = ws.max_row

        current_col = 1
        data = []
        num_dat = 0
        message_id = 0
        channel_id = 0

        async def button_callback1(interaction):
            
            user = interaction.user
            guild = interaction.guild
            member = await guild.fetch_member(user.id)
            if not member.guild_permissions.administrator:
                await interaction.response.send_message(content='진행자만 진행이 가능합니다.',ephemeral=True)
                return
                error_msg = f"{interaction.user.name}님이 다음버튼으로 진행함"
                await interaction.response.send_message(content=error_msg, delete_after=2)
            emoji = None
            nonlocal current_col, data, num_dat, message_id, channel_id
            
            if  num_dat>=0:
                channel = self.bot.get_channel(channel_id)
                message = await channel.fetch_message(message_id)
                await message.edit(view=None)
            cell = ws.cell(row=current_col, column=1).value
            channel_cell = ws.cell(row=current_col, column=2).value  # Get channel name from column 2
            data.append(cell)
            current_col += 1
            if not cell:
                channel = self.bot.get_channel(channel_id)
                message = await channel.fetch_message(message_id)
                await message.edit(view=None)
                await ctx.send(content='로그 종료')
                self.bot.all_p = '라운드1'
                return
            if channel_cell:  # if channel name exists
                channel = discord.utils.get(ctx.guild.channels, name=channel_cell)  # find channel by name
                if channel:
                    if data[num_dat].startswith('http'):
                        a = await channel.send(content=str(data[num_dat]))  # send message to found channel
                    elif not data[num_dat].startswith('http'):
                        embed = Embed(title=None, description=(data[num_dat]))
                        a = await channel.send(embed=embed)
                    message_id = a.id
                    channel_id = channel.id
                    num_dat += 1
                else:
                    await ctx.send('채널명이 맞는지 확인해주세요')  # send error message if channel not found
            else:
                if data[num_dat].startswith('http'):
                    a = await ctx.send(content=data[num_dat])
                elif not data[num_dat].startswith('http'):
                    embed = Embed(title=None, description=(data[num_dat]))
                    a = await ctx.send(embed=embed)
                message_id = a.id
                channel_id = ctx.channel.id
                num_dat += 1
            button1 = Button(label="다음", style=discord.ButtonStyle.green, disabled=False)
            button1.callback = button_callback1
            view = View(timeout=None)
            view.add_item(button1)
            await a.edit(view=view)

        button1 = Button(label="다음", style=discord.ButtonStyle.green, disabled=False)
        button1.callback = button_callback1

        view = View(timeout=None)
        view.add_item(button1)

        a = await ctx.channel.send(content=f'{file_name}',view=view)
        message_id = a.id
        channel_id = ctx.channel.id


    # @commands.command(name='로그1')
    # async def cmd_log(self, ctx: commands.Context,*args):
    #     a=args[0:]
    #     file_name=' '.join(s for s in a)
    #     await self.read_log(ctx, file_name)

    @commands.hybrid_command(name='로그1', with_app_command=True, description="엑셀로 저장된 로그를 불러옵니다.관리자만누름가능")
    @commands.guild_only()
    async def cmd_log_with_app_command(self, ctx: commands.Context, *,file_name: str):
        if hasattr(ctx, 'interaction') and ctx.interaction is not None:
            await ctx.send(content='https://i.imgur.com/gk3iHuX.gif',delete_after=0.00001)
        await self.read_log(ctx, file_name)

async def setup(bot):
    await bot.add_cog(Log1(bot), guilds=None)  # Add guild IDs if needed
