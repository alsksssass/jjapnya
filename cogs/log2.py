import discord
from discord import ui
from discord.ui import Button, View
from discord.ext import commands, tasks
from discord.ext.commands import has_permissions, CheckFailure, HybridCommand
import time
import os
import pymongo
from datetime import datetime
import asyncio
from asyncio import sleep
import json 
import socket
import discord.utils
import requests
from pydub import AudioSegment
import urllib.request
import re
from urllib.parse import urlparse
from discord import Embed
import openpyxl



class Log_t(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

    async def read_log(self, ctx: commands.Context, file_name: str):
            if not ctx.author.guild_permissions.administrator:
                await ctx.channel.send('권한이 없습니다.', delete_after=1)
                return
            time_sc=1
            folder_path = f"./dat/{ctx.guild.id}"
            files = os.listdir(folder_path)

            new_file_name = None
            for file in files:
                name, extension = os.path.splitext(file)
                if name == file_name:
                    new_file_name = name + extension
                    break
            if new_file_name is None:
                await ctx.send(f'파일 {file_name}이 존재하지 않습니다.\n!파일보기 로 확인하세요')
                return
            
            wb = openpyxl.load_workbook(f'./dat/{ctx.guild.id}/{new_file_name}')
            ws = wb.worksheets[0]
            sheet = wb.worksheets[0]

            max_column = sheet.max_column
            max_row = ws.max_row

            current_col = 1

            num_dat = 0


            
            # if (ws.cell(row=current_col, column=3).value)is None:
            #     await ctx.send('엑셀에 타임설정이 되어있지 않습니다.')
            #     return

            emoji = None

            a = await ctx.channel.send(content=f'{file_name}')


            cell = ws.cell(row=current_col, column=1).value
            # channel_cell = ws.cell(row=current_col, column=2).value  # Get channel name from column 2


            while current_col < max_row:
                ex_time = ws.cell(row=current_col, column=3).value
                if ex_time:
                    d_time=int(ex_time)*1
                else:
                    d_time=0
                channel_cell = ws.cell(row=current_col, column=2).value  # Get channel name from column 2
                if channel_cell:  # if channel name exists
                    await asyncio.sleep(d_time)
                    channel = discord.utils.get(ctx.guild.channels, name=channel_cell)  # find channel by name
                    if channel:
                        if ws.cell(row=current_col, column=1).value.startswith('http'):
                            a = await channel.send(content=str(ws.cell(row=current_col, column=1).value))  # send message to found channel
                        elif not ws.cell(row=current_col, column=1).value.startswith('http'):
                            embed = Embed(title=None, description=(ws.cell(row=current_col, column=1).value))
                            a = await channel.send(embed=embed)
                        num_dat += 1
                        current_col += 1
                        
                    else:
                        await ctx.send('채널명이 맞는지 확인해주세요')
                        return# send error message if channel not found
                else:
                    await asyncio.sleep(d_time)
                    if ws.cell(row=current_col, column=1).value.startswith('http'):
                        a = await ctx.send(content=ws.cell(row=current_col, column=1).value)
                    elif not ws.cell(row=current_col, column=1).value.startswith('http'):
                        embed = Embed(title=None, description=(ws.cell(row=current_col, column=1).value))
                        a = await ctx.send(embed=embed)
                    num_dat += 1
                    current_col += 1
                    

            await ctx.send(content='로그 종료')


    # @commands.command(name='로그타임')
    # async def cmd_log(self, ctx: commands.Context,time_sc:str, file_name:str):
    #     # a=args[0:]
    #     # file_name=' '.join(s for s in a)
    #     await self.read_log(ctx,time_sc,file_name)

    @commands.hybrid_command(name='로그타임', with_app_command=True, description="엑셀로 저장된 로그를 지정한 시간만큼 지연해서 불러옵니다.")
    @commands.guild_only()
    async def cmd_log_with_app_command(self, ctx: commands.Context, file_name: str):
        if hasattr(ctx, 'interaction') and ctx.interaction is not None:
            await ctx.send(content='https://i.imgur.com/gk3iHuX.gif',delete_after=0.00001)
        await self.read_log(ctx, file_name)

async def setup(bot):
    await bot.add_cog(Log_t(bot), guilds=None)  # Add guild IDs if needed
