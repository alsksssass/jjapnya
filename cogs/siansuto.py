import discord
from discord.ext import commands
from discord.ui import Button, View
from openpyxl import load_workbook

class MyCog(commands.Cog):
    def __init__(self, bot):
        self.bot = bot
        self.category_id = {}  # 카테고리별 메시지 아이디 저장용 딕셔너리
        self.message_buttons = {}  # 메시지 아이디와 버튼들의 매핑을 저장하는 딕셔너리
        self.clue_r = {}
        self.click_user_list = {}  # 버튼을 누른 유저의 닉네임을 저장하는 리스트
        self.output={}
        self.pre_button={}
        self.c_clue_r={}
        self.row={}
        
    def get_value_by_key(self,ctx, key):
        return self.category_id[str(ctx.guild.id)].get(key)
    async def get_emoji_id(self,ctx, emoji_name: str):
        for emoji in self.bot.emojis:
            if emoji.name == emoji_name:
                return emoji.id
    @commands.hybrid_command ( name = '버튼', with_app_command = True,description='버튼 폼 호출' )
    @commands.guild_only()
    async def 버튼_s(self, ctx, order, num: int = 0):
            if hasattr(ctx, 'interaction') and ctx.interaction is not None:
                await ctx.send(content='https://i.imgur.com/gk3iHuX.gif',delete_after=0.00001)
            if ctx.author.guild_permissions.administrator:
                if str(ctx.guild.id) not in self.pre_button:
                    self.pre_button[str(ctx.guild.id)]=[]
                pre_button = self.pre_button[str(ctx.guild.id)]
                if str(ctx.guild.id) not in self.click_user_list:
                    self.click_user_list[str(ctx.guild.id)]=[]
                if str(ctx.guild.id) not in self.message_buttons:
                    self.message_buttons[str(ctx.guild.id)]={}
                    self.category_id[str(ctx.guild.id)]={}
                    self.c_clue_r[str(ctx.guild.id)]=[]
                    # self.c_clue_r[str(ctx.guild.id)].append(ctx.channel.id)
                    print(self.c_clue_r[str(ctx.guild.id)])
                wb = load_workbook(f'./dat/{ctx.guild.id}/{order}.xlsx')
                ws = wb.active
                guild_id = 975632483750137878  # Replace with the ID of the guild (server)
                guild = self.bot.get_guild(guild_id)  # bot is your Discord bot instance
                emoji_id = await self.get_emoji_id(ctx,emoji_name='cat_S')  # Replace with the ID of the emoji
                emoji = guild.get_emoji(emoji_id)
                headers = [cell.value for cell in ws[1]]
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                if str(ctx.guild.id) not in self.row:
                    self.row[str(ctx.guild.id)] = len(rows)
                else:
                    self.row[str(ctx.guild.id)] = self.row[str(ctx.guild.id)] + (len(rows))
                self.clue_r[str(ctx.guild.id)] = num
                aa= await ctx.channel.send(f'남은 총 단서{self.row[str(ctx.guild.id)]}개, \n열람횟수{self.clue_r[str(ctx.guild.id)]}개 남음')
                self.c_clue_r[str(ctx.guild.id)].append(str(ctx.channel.id))
                self.c_clue_r[str(ctx.guild.id)].append(str(aa.id))
                await aa.add_reaction(emoji)
                print(self.c_clue_r[str(ctx.guild.id)])
                order_dict = {}

                async def button_callback(interaction, row):  # row를 추가 파라미터로 사용
                    
                    button_id = interaction.data['custom_id']
                    if 'pre_button' in headers and row[headers.index('pre_button')]:
                        if not row[headers.index('pre_button')] in pre_button:
                            await interaction.response.send_message("사전조사가 필요한 단서입니다.", delete_after=1)
                            return
                        
                        
                    if self.clue_r[str(ctx.guild.id)] > 0:
                        pre_button.append(button_id)

                        category, title = button_id.split('_')
                        if 'alive' in headers and row[headers.index('alive')] is False:
                            if category not in order_dict or title not in [button.label for button in order_dict[category].children]:
                                await interaction.response.send_message("유효하지 않은 버튼 ID입니다.", delete_after=1)
                                return
                        if 'permisson' in headers and row[headers.index('permisson')]:
                            permission_names = row[headers.index('permisson')].split(',')
                            member = ctx.guild.get_member(interaction.user.id)
                            for role_name in permission_names:
                                role = discord.utils.get(guild.roles, name=role_name)
                                if role:
                                    await member.add_roles(role)
                                else:
                                    await interaction.response.send_message(f'{role_name}이 없습니다')
                        if 'channel_permisson' in headers and row[headers.index('channel_permisson')]:
                            channel_names = row[headers.index('channel_permisson')].split(',')
                            member = ctx.guild.get_member(interaction.user.id)
                            for channel_name in channel_names:
                                channel = discord.utils.get(ctx.guild.channels, name=channel_name)
                                if channel:
                                    overwrite = discord.PermissionOverwrite()
                                    overwrite.connect=True
                                    overwrite.read_messages=True
                                    
                                    await channel.set_permissions(member, overwrite=overwrite)
                                else:
                                    await interaction.response.send_message(f'{channel_name}채널이름이 없습니다.')
                        view = order_dict[category]
                        for button in view.children:
                            if button.custom_id == button_id:
                                if 'alive' in headers and row[headers.index('alive')]:
                                    if 'info' in headers and row[headers.index('info')]:
                                        pass
                                    else:
                                        button.label=f'{button.label}->{interaction.user.nick}'
                                    pass
                                else:
                                    button.disabled = True
                                    if 'info' in headers and row[headers.index('info')]:
                                        pass
                                    else:
                                        button.label=f'{button.label}->{interaction.user.nick}'
                                
                        category_order = f"{category}_{order}"  # 카테고리별 메시지 아이디 저장용 딕셔너리에 사용할 키
                        category_message_id = self.get_value_by_key(ctx=ctx,key=category_order)
                        if category_message_id:
                            try:
                                category_message = await ctx.channel.fetch_message(category_message_id)
                                await category_message.edit(content=category, view=order_dict[category])
                            except discord.NotFound:
                                pass
                        if 'chance' in headers and row[headers.index('chance')]:
                            data1=str(row[headers.index('chance')])
                            if data1.isdigit():
                                num1=int(row[headers.index('chance')])
                                self.clue_r[str(ctx.guild.id)]+=num1
                                await ctx.send(f'조사권이 {num1}개 추가되었습니다.', delete_after=1)
                            else:
                                await ctx.send(f'{data1}은 숫자형식이 아닙니다.\n조사권이 추가되지 않았습니다.', delete_after=3)
                        await interaction.response.send_message(f"{(self.clue_r[str(ctx.guild.id)]-1)}개 남았습니다.", delete_after=1)

                        embed = discord.Embed(title=button_id)
                        image_descriptions = []
                        if 'description1' in headers and row[headers.index('description1')]:
                            for col in headers:
                                if col.startswith('description'):
                                    description = row[headers.index(col)]
                                    if description:
                                        if description.startswith('http'):
                                            image_descriptions.append(description)
                                        else:
                                            embed.add_field(name='', value=description, inline=False)
                            if 'channel' in headers and row[headers.index('channel')]:
                                channel_name = str(row[headers.index('channel')])
                                channel = discord.utils.get(ctx.guild.text_channels, name=channel_name)
                                if channel is None:
                                    channel = discord.utils.get(ctx.guild.voice_channels, name=channel_name)
                                if channel is None:
                                    channel = discord.utils.get(ctx.guild.categories, name=channel_name)
                                
                                if channel:
                                    message = await channel.send(embed=embed)
                                    self.message_buttons[str(ctx.guild.id)][message.id] = button_id  # 메시지 아이디와 버튼 ID 매핑 저장
                                    if image_descriptions:
                                        for img in image_descriptions:
                                            await channel.send(img)
                                else:
                                    message = await ctx.send(embed=embed)
                                    self.message_buttons[str(ctx.guild.id)][message.id] = button_id  # 메시지 아이디와 버튼 ID 매핑 저장
                                    if image_descriptions:
                                        for img in image_descriptions:
                                            await ctx.send(img)
                            else:
                                message = await ctx.send(embed=embed)
                                self.message_buttons[str(ctx.guild.id)][message.id] = button_id  # 메시지 아이디와 버튼 ID 매핑 저장
                                if image_descriptions:
                                    for img in image_descriptions:
                                        await ctx.send(img)
                        else:
                            await ctx.send(f'{interaction.user.nick}님이 선택하셧습니다', delete_after=1)


                        self.clue_r[str(ctx.guild.id)] -= 1
                        if self.clue_r[str(ctx.guild.id)]==0:
                            self.bot.all_p = '라운드2'
                        click_user = interaction.user.nick
                        self.click_user_list[str(ctx.guild.id)].append(click_user)
                        user_counts = {user: self.click_user_list[str(ctx.guild.id)].count(user) for user in self.click_user_list[str(ctx.guild.id)]}
                        self.row[str(ctx.guild.id)]-=1
                        self.output[str(ctx.guild.id)] = '\n'.join([f"{user}: {count}번" for user, count in user_counts.items()])
                        for i in range(0, len(self.c_clue_r[str(ctx.guild.id)]), 2):
                                # 짝수 인덱스는 채널 값
                            try:
                                channel_id = self.c_clue_r[str(ctx.guild.id)][i]
                                channel = self.bot.get_channel(int(channel_id))
                                
                                # 홀수 인덱스는 메시지 값
                                message_id = self.c_clue_r[str(ctx.guild.id)][i+1]
                                message = await channel.fetch_message(int(message_id))
                                
                                # 메시지를 수정
                                await message.edit(content=f'남은 총 단서{self.row[str(ctx.guild.id)]}개, \n열람횟수{self.clue_r[str(ctx.guild.id)]}개 남음\n{self.output[str(ctx.guild.id)]}')
                            except Exception as e:
                                print(e)
                                del self.c_clue_r[str(ctx.guild.id)][i:i+2]
                                pass
                        # for i in range(len(self.c_clue_r[str(ctx.guild.id)])):

                        #         channel = self.bot.get_channel(int(self.c_clue_r[str(ctx.guild.id)][0]))
                        #         if i!=0:
                        #             message = await channel.fetch_message(int(self.c_clue_r[str(ctx.guild.id)][i]))
                        #             await message.edit(content=f'남은 총 단서{self.row[str(ctx.guild.id)]}개, \n열람횟수{self.clue_r[str(ctx.guild.id)]}개 남음\n{self.output[str(ctx.guild.id)]}')
                    else:
                        try:
                            await interaction.response.send_message("추가 조사권이 없습니다.", delete_after=1)
                        except:
                            await ctx.send("추가 조사권이 없습니다.", delete_after=1)
                
                for row in rows:
                    category = row[headers.index('category')]
                    title = row[headers.index('title')]
                    custom_id = f'{category}_{title}'

                    view = order_dict.get(category)
                    if not view:
                        view = View(timeout=None)
                        order_dict[category] = view

                    button = Button(label=title, custom_id=custom_id, style=discord.ButtonStyle.green)
                    if button.label and button.label.strip():  # 버튼 라벨이 비어 있지 않은 경우에만 추가
                        button.callback = lambda i, r=row: button_callback(i, r)  # 람다 함수를 사용하여 올바른 row 데이터를 전달
                        view.add_item(button)
                for category, view in order_dict.items():
                    message = await ctx.send(category, view=view)
                    category_order = f"{category}_{order}"  # 카테고리별 메시지 아이디 저장용 딕셔너리에 사용할 키
                    self.category_id[str(ctx.guild.id)][category_order] = message.id  # 카테고리와 메시지 아이디 추가
                    
    @commands.hybrid_command ( name = '버튼체크', with_app_command = True,description='버튼 엑셀의 잘못된 부분을 표기해줍니다.' )
    @commands.guild_only()
    async def 버튼_s1(self, ctx, order, num: int = 0):
        if hasattr(ctx, 'interaction') and ctx.interaction is not None:
            await ctx.send(content='https://i.imgur.com/gk3iHuX.gif',delete_after=0.00001)
        if ctx.author.guild_permissions.administrator:
            if str(ctx.guild.id) not in self.pre_button:
                self.pre_button[str(ctx.guild.id)]=[]
            pre_button = self.pre_button[str(ctx.guild.id)]
            if str(ctx.guild.id) not in self.click_user_list:
                self.click_user_list[str(ctx.guild.id)]=[]
            if str(ctx.guild.id) not in self.message_buttons:
                self.message_buttons[str(ctx.guild.id)]={}
                self.category_id[str(ctx.guild.id)]={}
                self.c_clue_r[str(ctx.guild.id)]=[]
                self.c_clue_r[str(ctx.guild.id)].append(ctx.channel.id)
                print(self.c_clue_r[str(ctx.guild.id)])
            wb = load_workbook(f'./dat/{ctx.guild.id}/{order}.xlsx')
            ws = wb.active
            guild_id = 975632483750137878  # Replace with the ID of the guild (server)
            guild = self.bot.get_guild(guild_id)  # bot is your Discord bot instance
            emoji_id = await self.get_emoji_id(ctx,emoji_name='cat_S')  # Replace with the ID of the emoji
            emoji = guild.get_emoji(emoji_id)
            headers = [cell.value for cell in ws[1]]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            self.clue_r[str(ctx.guild.id)] = num
            aa= await ctx.channel.send(f'{self.clue_r[str(ctx.guild.id)]}개 남음')
            self.c_clue_r[str(ctx.guild.id)].append(aa.id)
            await aa.add_reaction(emoji)
            order_dict = {}

            async def button_callback(interaction, row):  # row를 추가 파라미터로 사용
                button_id = interaction.data['custom_id']
                if 'pre_button' in headers and row[headers.index('pre_button')]:
                    if not row[headers.index('pre_button')] in pre_button:
                        await interaction.response.send_message("사전조사가 필요한 단서입니다.", delete_after=1)
                        return
                    
                    
                if self.clue_r[str(ctx.guild.id)] > 0:
                    pre_button.append(button_id)

                    category, title = button_id.split('_')
                    if 'alive' in headers and row[headers.index('alive')] is False:
                        if category not in order_dict or title not in [button.label for button in order_dict[category].children]:
                            await interaction.response.send_message("유효하지 않은 버튼 ID입니다.", delete_after=1)
                            return
                    if 'permisson' in headers and row[headers.index('permisson')]:
                        permission_names = row[headers.index('permisson')].split(',')
                        member = ctx.guild.get_member(interaction.user.id)
                        for role_name in permission_names:
                            role = discord.utils.get(guild.roles, name=role_name)
                            if role:
                                await member.add_roles(role)
                            else:
                                await interaction.response.send_message(f'{role_name}이 없습니다')
                    if 'channel_permisson' in headers and row[headers.index('channel_permisson')]:
                        channel_names = row[headers.index('channel_permisson')].split(',')
                        member = ctx.guild.get_member(interaction.user.id)
                        for channel_name in channel_names:
                            channel = discord.utils.get(ctx.guild.channels, name=channel_name)
                            if channel:
                                overwrite = discord.PermissionOverwrite()
                                overwrite.connect=True
                                overwrite.read_messages=True
                                
                                await channel.set_permissions(member, overwrite=overwrite)
                            else:
                                await interaction.response.send_message(f'{channel_name}채널이름이 없습니다.')
                    view = order_dict[category]
                    for button in view.children:
                        if button.custom_id == button_id:
                            if 'alive' in headers and row[headers.index('alive')]:
                                if 'info' in headers and row[headers.index('info')]:
                                    pass
                                else:
                                    button.label=f'{button.label}->{interaction.user.nick}'
                                pass
                            else:
                                button.disabled = True
                                if 'info' in headers and row[headers.index('info')]:
                                    pass
                                else:
                                    button.label=f'{button.label}->{interaction.user.nick}'
                            
                    category_order = f"{category}_{order}"  # 카테고리별 메시지 아이디 저장용 딕셔너리에 사용할 키
                    category_message_id = self.get_value_by_key(ctx=ctx,key=category_order)
                    if category_message_id:
                        try:
                            category_message = await ctx.channel.fetch_message(category_message_id)
                            await category_message.edit(content=category, view=order_dict[category])
                        except discord.NotFound:
                            pass
                    if 'chance' in headers and row[headers.index('chance')]:
                        data1=str(row[headers.index('chance')])
                        if data1.isdigit():
                            num1=int(row[headers.index('chance')])
                            self.clue_r[str(ctx.guild.id)]+=num1
                            await ctx.send(f'조사권이 {num1}개 추가되었습니다.', delete_after=1)
                        else:
                            await ctx.send(f'{data1}은 숫자형식이 아닙니다.\n조사권이 추가되지 않았습니다.', delete_after=3)
                    await interaction.response.send_message(f"{(self.clue_r[str(ctx.guild.id)]-1)}개 남았습니다.", delete_after=1)

                    embed = discord.Embed(title=button_id)
                    image_descriptions = []
                    if 'description1' in headers and row[headers.index('description1')]:
                        for col in headers:
                            if col.startswith('description'):
                                description = row[headers.index(col)]
                                if description:
                                    if description.startswith('http'):
                                        image_descriptions.append(description)
                                    else:
                                        embed.add_field(name='', value=description, inline=False)
                        if 'channel' in headers and row[headers.index('channel')]:
                            channel_name = str(row[headers.index('channel')])
                            channel = discord.utils.get(ctx.guild.text_channels, name=channel_name)
                            if channel is None:
                                channel = discord.utils.get(ctx.guild.voice_channels, name=channel_name)
                            if channel is None:
                                channel = discord.utils.get(ctx.guild.categories, name=channel_name)
                            
                            if channel:
                                message = await channel.send(embed=embed)
                                self.message_buttons[str(ctx.guild.id)][message.id] = button_id  # 메시지 아이디와 버튼 ID 매핑 저장
                                if image_descriptions:
                                    for img in image_descriptions:
                                        await channel.send(img)
                            else:
                                message = await ctx.send(embed=embed)
                                self.message_buttons[str(ctx.guild.id)][message.id] = button_id  # 메시지 아이디와 버튼 ID 매핑 저장
                                if image_descriptions:
                                    for img in image_descriptions:
                                        await ctx.send(img)
                        else:
                            message = await ctx.send(embed=embed)
                            self.message_buttons[str(ctx.guild.id)][message.id] = button_id  # 메시지 아이디와 버튼 ID 매핑 저장
                            if image_descriptions:
                                for img in image_descriptions:
                                    await ctx.send(img)
                    else:
                        await ctx.send(f'{interaction.user.nick}님이 선택하셧습니다', delete_after=1)


                    self.clue_r[str(ctx.guild.id)] -= 1
                    if self.clue_r[str(ctx.guild.id)]==0:
                        self.bot.all_p = '라운드2'
                    click_user = interaction.user.nick
                    self.click_user_list[str(ctx.guild.id)].append(click_user)
                    user_counts = {user: self.click_user_list[str(ctx.guild.id)].count(user) for user in self.click_user_list[str(ctx.guild.id)]}
                    
                    self.output[str(ctx.guild.id)] = '\n'.join([f"{user}: {count}번" for user, count in user_counts.items()])
                    # for i in range(len(self.c_clue_r[str(ctx.guild.id)])):
                    #     channel = self.bot.get_channel(int(self.c_clue_r[str(ctx.guild.id)][0]))
                    #     if i!=0:
                    #         message = await channel.fetch_message(int(self.c_clue_r[str(ctx.guild.id)][i]))
                    #         await message.edit(content=f"{self.clue_r[str(ctx.guild.id)]}개 남음\n{self.output[str(ctx.guild.id)]}")
                    for i in range(0, len(self.c_clue_r[str(ctx.guild.id)]), 2):
                            # 짝수 인덱스는 채널 값
                        channel_id = self.c_clue_r[str(ctx.guild.id)][i]
                        channel = self.bot.get_channel(int(channel_id))
                        
                        # 홀수 인덱스는 메시지 값
                        message_id = self.c_clue_r[str(ctx.guild.id)][i+1]
                        message = await channel.fetch_message(int(message_id))
                        
                        # 메시지를 수정
                        await message.edit(content=f"{self.clue_r[str(ctx.guild.id)]}개 남음\n{self.output[str(ctx.guild.id)]}")

                    print(self.output[str(ctx.guild.id)])
                else:
                    try:
                        await interaction.response.send_message("추가 조사권이 없습니다.", delete_after=1)
                    except:
                        await ctx.send("추가 조사권이 없습니다.", delete_after=1)
            
            for row in rows:
                category = row[headers.index('category')]
                title = row[headers.index('title')]
                custom_id = f'{category}_{title}'

                view = order_dict.get(category)
                if not view:
                    view = View(timeout=None)
                    order_dict[category] = view

                button = Button(label=title, custom_id=custom_id, style=discord.ButtonStyle.green)
                if button.label and button.label.strip():  # 버튼 라벨이 비어 있지 않은 경우에만 추가
                    button.callback = lambda i, r=row: button_callback(i, r)  # 람다 함수를 사용하여 올바른 row 데이터를 전달
                    view.add_item(button)
                else:
                    for index, rowt in enumerate(rows):
                        if rowt == row:
                            await ctx.send(f'{(index+2)}번째 행의 데이터가 잘못입력 되어있습니다. 확인 해 주세요.\n{row}')
            for category, view in order_dict.items():
                message = await ctx.send(category, view=view)
                category_order = f"{category}_{order}"  # 카테고리별 메시지 아이디 저장용 딕셔너리에 사용할 키
                self.category_id[str(ctx.guild.id)][category_order] = message.id  # 카테고리와 메시지 아이디 추가
        
    @commands.hybrid_command ( name = '버튼추가', with_app_command = True,description='버튼 횟수 추가' )
    @commands.guild_only()
    async def add_button(self,ctx,num:int):
        if ctx.author.guild_permissions.administrator:
            self.clue_r[str(ctx.guild.id)] += num
            for i in range(0, len(self.c_clue_r[str(ctx.guild.id)]), 2):
                    # 짝수 인덱스는 채널 값
                try:
                    channel_id = self.c_clue_r[str(ctx.guild.id)][i]
                    channel = self.bot.get_channel(int(channel_id))
                    
                    # 홀수 인덱스는 메시지 값
                    message_id = self.c_clue_r[str(ctx.guild.id)][i+1]
                    message = await channel.fetch_message(int(message_id))
                    
                    # 메시지를 수정
                    await message.edit(content=f'남은 총 단서{self.row[str(ctx.guild.id)]}개, \n열람횟수{self.clue_r[str(ctx.guild.id)]}개 남음\n{self.output[str(ctx.guild.id)]}')
                except Exception as e:
                    print(e)
                    del self.c_clue_r[str(ctx.guild.id)][i:i+2]
                    pass
            await ctx.send(f'{num}개 추가되었습니다. 현재{self.clue_r[str(ctx.guild.id)]}개 남음 ',delete_after=5)
            
    @commands.hybrid_command ( name = '버튼초기화', with_app_command = True,description='버튼 초기화(버튼사용전 필수!!)' )
    @commands.guild_only()
    async def re_set(self,ctx):
        if ctx.author.guild_permissions.administrator:
            dicts_to_clear = [
                self.pre_button,
                self.message_buttons,
                self.click_user_list,
                self.category_id,
                self.clue_r,
                self.c_clue_r,
                self.output,
                self.row
            ]

            guild_id = str(ctx.guild.id)
            for d in dicts_to_clear:
                if guild_id in d:
                    if isinstance(d[guild_id], (int, str)):
                        d[guild_id] = 0  # 정수 값의 경우 0으로 초기화
                    else:
                        d[guild_id].clear()
            await ctx.send("초기화 되었습니다.", delete_after=1)
        
async def setup(bot):
    await bot.add_cog(MyCog(bot))
